##############################################################################
# ThanCad 0.6.7 "Students2021": n-dimensional CAD with raster support for engineers
#
# Copyright (C) 2001-2022 Thanasis Stamos, May 26, 2022
# Athens, Greece, Europe
# URL: http://thancad.sourceforge.net
# e-mail: cyberthanasis@gmx.net
#
# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details (www.gnu.org/licenses/gpl.html).
#
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  USA
##############################################################################
"""\
ThanCad 0.6.7 "Students2021": n-dimensional CAD with raster support for engineers

This module defines objects which read spreadsheet files (.xls, .xlsx, .ods)
for points, lines, texts or surface and create the appropriate ThanCad's elements.
"""

try: import xlrd          #Hopefully py2exe and cxFreeze get the hint to include xlrd
except ImportError: pass
try: import openpyxl      #Hopefully py2exe and cxFreeze get the hint to include openpyxl
except ImportError: pass
from p_ggen import Tgui as T


class ThanXlsReader:
    "A class to read point coordinates from a spreadsheet (xls) file using xlrd."

    @staticmethod
    def isEmpty(sh, i, j):
        "Test if cell is empty."
        if sh.cell_type(i, j) == xlrd.XL_CELL_EMPTY: return True  #Empty cell
        if sh.cell_type(i, j) == xlrd.XL_CELL_BLANK: return True  #Blank text
        if sh.cell_type(i, j) == xlrd.XL_CELL_TEXT:
            return sh.cell_value(i, j).strip() == ""              #Blank text
        return False                                              #Number, date, boolean, error


    @staticmethod
    def toString(sh, i, j):
        "Transform content of cell to string."
        if sh.cell_type(i, j) == xlrd.XL_CELL_EMPTY: return ""  #Empty cell
        if sh.cell_type(i, j) == xlrd.XL_CELL_BLANK: return ""  #Blank text
        if sh.cell_type(i, j) == xlrd.XL_CELL_TEXT: return sh.cell_value(i, j) #text
        v = sh.cell_value(i, j)
        if v == int(v): v = int(v)   #Αν τα δεκαδικά είναι 0, είναι κατά 99.99% σίγουρο ότι είναι ακέραιος
        return str(v)                #Convert number to string


    @staticmethod
    def toFloat(sh, i, j):
        "Transform content of cell to real number."
        if sh.cell_type(i, j) == xlrd.XL_CELL_EMPTY: return 0.0  #Empty cell
        if sh.cell_type(i, j) == xlrd.XL_CELL_BLANK: return 0.0  #Blank text
        if sh.cell_type(i, j) == xlrd.XL_CELL_TEXT:
            t = sh.cell_value(i, j)
            if t.strip() == "": return 0.0                       #Blank text (may contain white characters)
            try: return float(t.replace(",", "."))               #Convert text to number
            except: return None                                  #Conversion failed
        if sh.cell_type(i, j) == xlrd.XL_CELL_ERROR: return None #Error code
        if sh.cell_type(i, j) == xlrd.XL_CELL_DATE:  return None #Date
        try:
            return float(sh.cell_value(i, j))                    #Try to convert int, boolean, float or other to float
        except Exception:
            return None                                          #Failed

    @staticmethod
    def toInt(sh, i, j):
        "Transform content of cell to integer number."
        return toIntxx(ThanXlsReader, sh, i, j)

    @staticmethod
    def openSpread(fn, minrows, mincols, prt, er2s):
        "Open the .xls file, make some tests and return the book object."
        try:
            import xlrd
        except ImportError:
            er2s(T["Can not import xls spreadsheets: The xlrd library/package was not found.\n"\
                "Please install xlrd in your system and retry."])  #Raises ThanImportError
        prt("Opening {}...".format(fn), "info1")
        try:
            book = xlrd.open_workbook(fn)
        except BaseException as e:
            er2s(str(e))                                           #Raises ThanImportError
        prt(T["The number of {0} worksheets is {1}"].format(fn, book.nsheets), "info1")
        if book.nsheets > 1: prt(T["Warning: Only the first sheet will be read"], "can")
        sh = ThanXlsReader.getSheet(book, 0)

        prt(T["Worksheet {0}: {1} rows, {2} columns"].format(sh.name, sh.nrows, sh.ncols), "info1")
        if sh.ncols < 1 or sh.nrows < 1:
            er2s(T["Sheet {} is empty!"].format(sh.name))
        elif sh.ncols < mincols:
            er2s(T["Sheet {} has less than {} columns!"].format(sh.name, mincols))
        elif sh.nrows < minrows:
            er2s(T["Sheet {} has less than {} rows!"].format(sh.name, minrows))
        return book, sh


    @staticmethod
    def getSheet(book, i):
        "Get the ith sheet of the book."
        if i >= book.nsheets: return None   #sheet does not exist
        sh = book.sheet_by_index(i)
        return sh


    @staticmethod
    def getNsheets(book):
        "Get the number of sheets of the book."
        return book.nsheets

    @staticmethod
    def closeSpread(book, sh):
        "Close that spreadsheet."
        #It seems that xlrd automatically closes the file (if the on_demand arguments is not True)
        #release_resources() is needed when on_demand=True, but ot does not hurt to call it.
        book.release_resources()


class ThanXlsxReader:
    "A class to read point coordinates from a spreadsheet (xlsx) file using openpyxl."

    @staticmethod
    def isEmpty(sh, i, j):
        "Test if cell is empty; i, j are zero based, but openpyxl is one based."
        cell1 = sh.cell(row=i+1, column=j+1)
        v = cell1.value
        if v is None: return True   #Empty cell
        if isinstance(v, str):
            return v.strip() == ""  #Blank text
        return False   #int, float, datetime.datetime, bool, nonblank text (error is string: '#VALUE!')


    @staticmethod
    def toString(sh, i, j):
        "Transform content of cell to string; i, j are zero based, but openpyxl is one based."
        cell1 = sh.cell(row=i+1, column=j+1)
        v = cell1.value
        if v is None: return ""      #Empty cell
        return str(v)                #Convert int, float, datetime.datetime, bool to string.
                                     #If it is string it remains unchanged


    @staticmethod
    def toFloat(sh, i, j):
        "Transform content of cell to real number; i, j are zero based, but openpyxl is one based."
        cell1 = sh.cell(row=i+1, column=j+1)
        v = cell1.value
        if v is None: return 0.0  #Empty cell
        if isinstance(v, str):
            if v.strip() == "": return 0.0                       #Blank text (may contain white characters)
            try: return float(v.replace(",", "."))               #Convert text to number
            except: return None                                  #Conversion failed
        try:
            return float(v)                                      #Try to convert int, boolean, float or other to float
        except Exception:
            return None                                          #Failed

    @staticmethod
    def toInt(sh, i, j):
        "Transform content of cell to integer number."
        return toIntxx(ThanXlsxReader, sh, i, j)

    @staticmethod
    def openSpread(fn, minrows, mincols, prt, er2s):
        "Open the .xlsx file, make some tests and return the book object."
        try:
            import openpyxl
        except ImportError:
            er2s(T["Can not import xlsx spreadsheets: The openpyxl library/package was not found.\n"\
                "Please install openpyxl in your system and retry."])  #Raises ThanImportError
        prt("Opening {}...".format(fn), "info1")
        try:
            book = openpyxl.load_workbook(filename=fn, data_only=True)
        except BaseException as e:
            er2s(str(e))                                               #Raises ThanImportError
        prt(T["The number of {0} worksheets is {1}"].format(fn, len(book.sheetnames)), "info1")
        if len(book.sheetnames) > 1: prt(T["Warning: Only the first sheet will be read"], "can")
        sh = ThanXlsxReader.getSheet(book, 0)

        prt(T["Worksheet {0}: {1} rows, {2} columns"].format(sh.name, sh.nrows, sh.ncols), "info1")
        if sh.ncols < 1 or sh.nrows < 1:
            er2s(T["Sheet {} is empty!"].format(sh.name))
        elif sh.ncols < mincols:
            er2s(T["Sheet {} has less than {} columns!"].format(sh.name, mincols))
        elif sh.nrows < minrows:
            er2s(T["Sheet {} has less than {} rows!"].format(sh.name, minrows))
        return book, sh


    @staticmethod
    def getSheet(book, i):
        "Get the ith sheet of the book."
        if i >= len(book.sheetnames): return None   #sheet does not exist
        sh = book.worksheets[i]
        sh.name = sh.title       #New object variable for compatibility with xlrd (does not exist in openpyxl ws)
        sh.nrows = sh.max_row    #New object variable for compatibility with xlrd (does not exist in openpyxl ws)
        sh.ncols = sh.max_column #New object variable for compatibility with xlrd (does not exist in openpyxl ws)
        return sh


    @staticmethod
    def getNsheets(book):
        "Get the number of sheets of the book."
        return len(book.sheetnames)


    @staticmethod
    def closeSpread(book, sh):
        "Close that spreadsheet."
        #Closes the workbook file if open. Only affects read-only and write-only modes.
        #It does not hurt to close it.
        book.close()


def toIntxx(rd, sh, i, j):
    "Transform content of cell to integer number."
    temp = rd.toFloat(sh, i, j)
    if temp is None: return temp
    try:    k = int(temp)
    except: return None
    if k != temp: return None
    return k
