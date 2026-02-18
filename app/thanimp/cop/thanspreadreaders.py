##############################################################################
# ThanCad 0.6.7 "Students2021": n-dimensional CAD with raster support for engineers
# 
# Copyright (C) 2001-2022 Thanasis Stamos, May 26, 2022
# Athens, Greece, Europe
# URL: http://thancad.sourceforge.net
# e-mail: cyberthanasis@gmx.net
# 
# This program is free software; you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation; either version 2 of the License, or
# (at your option) any later version.
# 
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details (www.gnu.org/licenses/gpl.html).
# 
# You should have received a copy of the GNU General Public License
# along with this program; if not, write to the Free Software
# Foundation, Inc., 59 Temple Place, Suite 330, Boston, MA  02111-1307  USA
##############################################################################
"""\
ThanCad 0.6.7 "Students2021": n-dimensional CAD with raster support for engineers

This module defines objects which read spreadsheet files (.xls, .xlsx, .ods)
for points, lines, texts or surface and create the appropriate ThanCad's elements.
"""

try: import xlrd          #Hopefully py2exe and cxFreeze get the hint to include xlrd
except ImportError: pass
try: import openpyxl      #Hopefully py2exe and cxFreeze get the hint to include openpyxl
except ImportError: pass
from thantrans import T


class ThanXlsReader:
    "A class to read point coordinates from a spreadsheet (xls) file using xlrd."

    @staticmethod
    def isEmpty(sh, i, j):
        "Test if cell is empty."
        if sh.cell_type(i, j) == xlrd.XL_CELL_EMPTY: return True  #Empty cell
        if sh.cell_type(i, j) == xlrd.XL_CELL_BLANK: return True  #Blank text
        if sh.cell_type(i, j) == xlrd.XL_CELL_TEXT:
            return sh.cell_value(i, j).strip() == ""              #Blank text
        return False                                              #Number, date, boolean, error


    @staticmethod
    def toString(sh, i, j):
        "Transform content of cell to string."
        if sh.cell_type(i, j) == xlrd.XL_CELL_EMPTY: return ""  #Empty cell
        if sh.cell_type(i, j) == xlrd.XL_CELL_BLANK: return ""  #Blank text
        if sh.cell_type(i, j) == xlrd.XL_CELL_TEXT: return sh.cell_value(i, j) #text
        v = sh.cell_value(i, j)
        if v == int(v): v = int(v)   #Αν τα δεκαδικά είναι 0, είναι κατά 99.99% σίγουρο ότι είναι ακέραιος
        return str(v)                #Convert number to string


    @staticmethod
    def toFloat(sh, i, j):
        "Transform content of cell to real number."
        if sh.cell_type(i, j) == xlrd.XL_CELL_EMPTY: return 0.0  #Empty cell
        if sh.cell_type(i, j) == xlrd.XL_CELL_BLANK: return 0.0  #Blank text
        if sh.cell_type(i, j) == xlrd.XL_CELL_TEXT:
            t = sh.cell_value(i, j)
            if t.strip() == "": return 0.0                       #Blank text (may contain white characters)
            try: return float(t.replace(",", "."))               #Convert text to number
            except: return None                                  #Conversion failed
        if sh.cell_type(i, j) == xlrd.XL_CELL_ERROR: return None #Error code
        if sh.cell_type(i, j) == xlrd.XL_CELL_DATE:  return None #Date
        try:
            return float(sh.cell_value(i, j))                    #Try to convert int, boolean, float or other to float
        except Exception:
            return None                                          #Failed


    @staticmethod
    def openSpread(fn, minrows, mincols, prt, er2s):
        "Open the .xls file, make some tests and return the book object."
        try:
            import xlrd
        except ImportError:
            er2s(T["Can not import xls spreadsheets: The xlrd library/package was not found.\n"\
                "Please install xlrd in your system and retry."])  #Raises ThanImportError
        prt("Opening {}...".format(fn), "info1")
        try:
            book = xlrd.open_workbook(fn)
        except BaseException as e:
            er2s(str(e))                                           #Raises ThanImportError
        prt(T["The number of {0} worksheets is {1}"].format(fn, book.nsheets), "info1")
        if book.nsheets > 1: prt(T["Warning: Only the first sheet will be read"], "can")
        sh = book.sheet_by_index(0)
        prt(T["Worksheet {0}: {1} rows, {2} columns"].format(sh.name, sh.nrows, sh.ncols), "info1")
        if sh.ncols < 1 or sh.nrows < 1:
            er2s(T["Sheet {} is empty!"].format(sh.name))
        elif sh.ncols < mincols:
            er2s(T["Sheet {} has less than {} columns!"].format(sh.name, mincols))
        elif sh.nrows < minrows:
            er2s(T["Sheet {} has less than {} rows!"].format(sh.name, minrows))
        return book, sh


    @staticmethod
    def closeSpread(book, sh):
        "Close that spreadsheet."
        #It seems that xlrd automatically closes the file (if the on_demand arguments is not True)
        #release_resources() is needed when on_demand=True, but ti does not hurt to call ir
        book.release_resources()


class ThanXlsxReader:
    "A class to read point coordinates from a spreadsheet (xlsx) file using openpyxl."

    @staticmethod
    def isEmpty(sh, i, j):
        "Test if cell is empty; i, j are zero based, but openpyxl is one based."
        cell1 = sh.cell(row=i+1, column=j+1)
        v = cell1.value
        if v is None: return True   #Empty cell
        if isinstance(v, str):
            return v.strip() == ""  #Blank text
        return False   #int, float, datetime.datetime, bool, nonblank text (error is string: '#VALUE!')


    @staticmethod
    def toString(sh, i, j):
        "Transform content of cell to string; i, j are zero based, but openpyxl is one based."
        cell1 = sh.cell(row=i+1, column=j+1)
        v = cell1.value
        if v is None: return ""      #Empty cell
        return str(v)                #Convert int, float, datetime.datetime, bool to string.
                                     #String remain uchanged


    @staticmethod
    def toFloat(sh, i, j):
        "Transform content of cell to real number; i, j are zero based, but openpyxl is one based."
        cell1 = sh.cell(row=i+1, column=j+1)
        v = cell1.value
        if v is None: return 0.0  #Empty cell
        if isinstance(v, str):
            if v.strip() == "": return 0.0                       #Blank text (may contain white characters)
            try: return float(v.replace(",", "."))               #Convert text to number
            except: return None                                  #Conversion failed
        try:
            return float(v)                                      #Try to convert int, boolean, float or other to float
        except Exception:
            return None                                          #Failed


    @staticmethod
    def openSpread(fn, minrows, mincols, prt, er2s):
        "Open the .xlsx file, make some tests and return the book object."
        try:
            import openpyxl
        except ImportError:
            er2s(T["Can not import xlsx spreadsheets: The openpyxl library/package was not found.\n"\
                "Please install openpyxl in your system and retry."])  #Raises ThanImportError
        prt("Opening {}...".format(fn), "info1")
        try:
            book = openpyxl.load_workbook(filename=fn, data_only=True)
        except BaseException as e:
            er2s(str(e))                                               #Raises ThanImportError
        prt(T["The number of {0} worksheets is {1}"].format(fn, len(book.sheetnames)), "info1")
        if len(book.sheetnames) > 1: prt(T["Warning: Only the first sheet will be read"], "can")
        sh = book.worksheets[0]

        sh.name = sh.title       #New object variable for compatibility with xlrd (does not exist in openpyxl ws)
        sh.nrows = sh.max_row    #New object variable for compatibility with xlrd (does not exist in openpyxl ws)
        sh.ncols = sh.max_column #New object variable for compatibility with xlrd (does not exist in openpyxl ws)

        prt(T["Worksheet {0}: {1} rows, {2} columns"].format(sh.name, sh.nrows, sh.ncols), "info1")
        if sh.ncols < 1 or sh.nrows < 1:
            er2s(T["Sheet {} is empty!"].format(sh.name))
        elif sh.ncols < mincols:
            er2s(T["Sheet {} has less than {} columns!"].format(sh.name, mincols))
        elif sh.nrows < minrows:
            er2s(T["Sheet {} has less than {} rows!"].format(sh.name, minrows))
        return book, sh


    @staticmethod
    def closeSpread(book, sh):
        "Close that spreadsheet."
        #Closes the workbook file if open. Only affects read-only and write-only modes.
        #It does not hurt to close it
        book.close()


class ThanImportSpread(ThanImportBase):
    "A mixin class to read point coordinates etc from a spreadsheet; xls/xlsx agnostic."

    def iteraxyz(self, sh, rd):
        """Iterate through all the point coordinates in the xls file: aa, x, y, z.

        icod is 0 if a point was found, 1 if empty line was found or 2 if syntax
        error or empty cell."""
        ztoo = sh.ncols > 3
        for i in range(sh.nrows):
            self._lindxf = i+1
            empty = rd.isEmpty(sh, i, 0) and rd.isEmpty(sh, i, 1) and rd.isEmpty(sh, i, 2)
            if ztoo: empty = empty and rd.isEmpty(sh, i, 3)
            if empty:
                yield 1, None, None, None, None
            else:
                aa = rd.toString(sh, i, 0)
                x = rd.toFloat(sh, i, 1)
                y = rd.toFloat(sh, i, 2)
                z = rd.toFloat(sh, i, 3) if ztoo else 0.0
                icod = 2 if x is None or y is None or z is None else 0
                yield icod, aa, x, y, z


    def itertexts(self, sh, rd):
        """Iterate through all the text definitions: x, y, size, theta, text.

        icod is 0 if a text was found, 1 if empty line was found or 2 if syntax
        error or empty cell."""
        for i in range(sh.nrows):
            self._lindxf = i+1
            empty = rd.isEmpty(sh, i, 0) and rd.isEmpty(sh, i, 1) and rd.isEmpty(sh, i, 2) \
                and rd.isEmpty(sh, i, 3) and rd.isEmpty(sh, i, 4)
            if empty:
                yield 1, None, None, None, None, None
            else:
                x = rd.toFloat(sh, i, 0)
                y = rd.toFloat(sh, i, 1)
                sizet = rd.toFloat(sh, i, 2)
                theta = rd.toFloat(sh, i, 3)
                text = rd.toString(sh, i, 4)
                icod = 2 if x is None or y is None or sizet is None or theta is None else 0
                yield icod, x, y, sizet, theta, text


    def iterxyz(self, sh, rd):
        """Iterate through all the point coordinates in the xls file: x, y, z.

        icod is 0 if a point was found, 1 if empty line was found or 2 if syntax
        error or empty cell."""
        ztoo = sh.ncols > 2
        for i in range(sh.nrows):
            self._lindxf = i+1
            empty = rd.isEmpty(sh, i, 0) and rd.isEmpty(sh, i, 1)
            if ztoo: empty = empty and rd.isEmpty(sh, i, 2)
            if empty:
                yield 1, None, None, None
            else:
                x = rd.toFloat(sh, i, 0)
                y = rd.toFloat(sh, i, 1)
                z = rd.toFloat(sh, i, 2) if ztoo else 0.0
                icod = 2 if x is None or y is None or z is None else 0
                yield icod, x, y, z


    def iterfloats(self, sh, rd):
        """Iterate through all lines of xls file and convert all cell values to floats.

        icod is 0 if valid floats were found in a line, 1 if an empty line was found or
        2 if syntax error or empty cell."""
        for i in range(sh.nrows):
            self._lindxf = i+1
            if all(rd.isEmpty(sh, i, j) for j in range(sh.ncols)): yield 1, None
            vs = [rd.toFloat(sh, i, j) for j in range(sh.ncols)]
            if any(v is None for v in vs): yield 2, vs
            yield 0, vs


    def thanImportPoints(self):
        "Import points from a spreadsheet."
        if self.fDxf.name.endswith(".xls"): rd = ThanXlsReader()
        else:                               rd = ThanXlsxReader()
        book, sh = rd.openSpread(self.fDxf.name, 1, 3, self.thanDr.prt, self.thanEr2s)
        np = 0
        handle = ""
        validc = [True, True, True]
        if sh.ncols < 4: validc[2] = False
        for icod, aa, x, y, z in self.iteraxyz(sh, rd):
            if icod == 1:
                self.thanWarn(T["blank row is ignored."])
            elif icod == 2:
                self.thanWarn(T["x, y or z is not a number. Point is ignored."])
            else:
                self.thanDr.dxfPoint(x, y, z, self.defLay, handle, None, aa, validc)
                np += 1
        lay = self.thanDr._dr.thanLayerTree.thanFindic(self.defLay)
        if lay is not None:   #If self.defLay != "0" and no points were added, self.defLay has not been created
            lay.thanAtts["hidename"].thanValSet(False)
            if sh.ncols > 3: lay.thanAtts["hideheight"].thanValSet(False)
        if sh.ncols > 3:
            self.thanDr.prt(T["{} points imported (out of {} spreadsheet rows)"].format(np, sh.nrows), "info")
        else:
            self.thanDr.prt(T["{} points (with no z) imported (out of {} spreadsheet rows)"].format(np, sh.nrows), "info")

        rd.closeSpread(book, sh)
        del sh, book


    def thanImportTexts(self):
        "Import texts from a spreadsheet."
        if self.fDxf.name.endswith(".xls"): rd = ThanXlsReader()
        else:                               rd = ThanXlsxReader()
        book, sh = rd.openSpread(self.fDxf.name, 1, 5, self.thanDr.prt, self.thanEr2s)
        np = 0
        handle = ""
        for icod, x, y, sizet, theta, text in self.itertexts(sh, rd):
            if icod == 1:
                self.thanWarn(T["blank row is ignored."])
            elif icod == 2:
                self.thanWarn(T["x, y, size or theta is not a number. Text is ignored."])
            else:
                self.thanDr.dxfText(x, y, 0.0, self.defLay, handle, None, text, sizet, theta)
                np += 1

        self.thanDr.prt(T["{} texts imported (out of {} spreadsheet rows)"].format(np, sh.nrows), "info")

        rd.closeSpread(book, sh)
        del sh, book


    def thanImportLines(self):
        "Import lines from a spreadsheet."
        if self.fDxf.name.endswith(".xls"): rd = ThanXlsReader()
        else:                               rd = ThanXlsxReader()
        book, sh = rd.openSpread(self.fDxf.name, 1, 2, self.thanDr.prt, self.thanEr2s)
        handle = ""
        xx = []
        yy = []
        zz = []
        for icod, x, y, z in self.iterxyz(sh, rd):
            if icod == 1:
                if len(xx) > 0:
                    self.thanDr.dxfPolyline(xx, yy, zz, self.defLay, handle, None)
                    xx = []
                    yy = []
                    zz = []
            elif icod == 2:
                self.thanEr1s(T["x, y or z is not a number."])
            else:
                xx.append(x)
                yy.append(y)
                zz.append(z)
        if len(xx) > 0:
            self.thanDr.dxfPolyline(xx, yy, zz, self.defLay, handle, None)

        rd.closeSpread(book, sh)
        del sh, book


    def thanImportSurface(self):
        """Import surface from a spreadsheet."

        The surface is defined as z values of a grid with possibly variable step.
        The first row contains the x value of each column, and the first column
        contains the y values of each row.
        The surface is imported as a set of neiboring 3dfaces.
        """

        if self.fDxf.name.endswith(".xls"): rd = ThanXlsReader()
        else:                               rd = ThanXlsxReader()
        book, sh = rd.openSpread(self.fDxf.name, 3, 3, self.thanDr.prt, self.thanEr2s)
        handle = ""
        #First line contains the x coordinates of all columns
        it = self.iterfloats(sh, rd)
        icod, xs = next(it)
        if icod == 1:
            self.thanEr1s(T["Blank row found."])
        elif icod == 2:
            if any(x is None for x in xs[1:]): #First cell may be blank or may contain anything
                self.thanEr1s(T["Some cells do not contain numbers."])
        #Second line contains the first set of z values
        icod, vsa = next(it)
        if icod == 1:
            self.thanEr1s(T["Blank row found."])
        elif icod == 2:
            self.thanEr1s(T["Some cells do not contain numbers."])
        #Iterate through 3rd line to end
        for icod, vsb in it:
            if icod == 1:
                self.thanEr1s(T["Blank row found."])
            elif icod == 2:
                self.thanEr1s(T["Some cells do not contain numbers."])
            ya = vsa[0]
            yb = vsb[0]
            for j in range(2, sh.ncols):
                xa = xs[j-1]
                xb = xs[j]
                xx = xa,       xb,     xb,     xa
                yy = ya,       ya,     yb,     yb
                zz = vsa[j-1], vsa[j], vsb[j], vsb[j-1]
                self.thanDr.dxf3dface(xx, yy, zz, self.defLay, handle, None)
            vsa = vsb

        rd.closeSpread(book, sh)
        del sh, book


class ThanImportXlsPoints(ThanImportSpread):
    "A class to import (topographic) points from a spreadsheet xls file."
    def thanImport(self): self.thanImportPoints()


class ThanImportXlsTexts(ThanImportSpread):
    "A class to import texts a spreadsheet (xls/xlsx) file."
    def thanImport(self): self.thanImportTexts()


class ThanImportXlsLines(ThanImportSpread):
    "A class to import lines from a spreadsheet (xls/xlsx) file."
    def thanImport(self): self.thanImportLines()


class ThanImportXlsSurface(ThanImportSpread):
    "A class to import a surface from a spreadsheet (xls/xlsx) file."
    def thanImport(self): self.thanImportSurface()
